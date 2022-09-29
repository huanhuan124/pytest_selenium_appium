import datetime
import json
#openpyxl模块是一个读写Excel 2010文档的Python库，如果要处理更早格式的Excel文档，需要用到额外的库，而且Excel中的坐标是从1，1开始，并不是0，0开始
import openpyxl
import requests

#资产登录
#预生产环境登录
def login_do():
    data={
    "loginId": "HYhLuyvEoBNfYk2Tt7u27Q==",
    "password": "r5wA9Eu7SML/wxmAewlekA=="
    }
    re=requests.post("http://**/login.do_",data=data)
    # print(re.headers)
    setcookie=re.headers['Set-Cookie'].split(";")[0]
    # print(re.headers['Set-Cookie'].split(";")[0])
    return setcookie

# setcookie = login_do()

#.新建准运营车辆
# http://**

frameNo=""
engineNo=""
# 预生产环境创建车辆
def createVehicle(frameNo,engineNo,setcookie):
    #由于直接抓取的接口参数是java的json，包含null类型，Python不认，需要先转成字符串（使用''''''方法），然后转成Python的json
    data ='''{	"baseInfoVo": {
            "vehicleId": "",
            "companyId": 46,
            "brandId": 67,
            "vehicleSeriesId": 560,
            "vehiclenoCityId": "",
            "vehiclenoCityName": "",
            "vehicleNo": "",
            "frameNo": "frameNo",
            "engineNo": "engineNo",
            "envStandard": 3,
            "modelId": 1221,
            "vehicleLicensemode": "No1238198",
            "vehicleModel300Name": "2015款 宝马5系GT(进口) 528i 领先型",
            "color": 3,
            "registerNo": "",
            "useNature": 0,
            "carUseType": 0,
            "vehicleRemark": "",
            "checkTime": "",
            "ownerStatus1st": 10,
            "ownerStatus2nd": 1001,
            "ownerStatus3rd": "",
            "contractNo": "llyHT002"	},	"operatingInfoVo": {
            "shortModelId": 1103,
            "shortModelName": "宝马BMW9/三厢/2.8自动",
            "businessType": 1,
            "isUCar": 0,
            "isHertz": 0,
            "isLease": 0,
            "onLineTime": "",
            "lastQuitRunTime": "",
            "handleTime": "",
            "ucarRentType": "",
            "oilVolume": null,
            "oilVolumeLatticeNum": 0,
            "orderCar": 0,
            "cityId": 1,
            "nowCityId": 1,
            "nowCityName": "北京",
            "parkCityId": 1,
            "departmentId": 10189,
            "nowDepartmentId": 10189,
            "parkDeptId": 10189,
            "runMilesInput": "3000",
            "nextInspecteTime": "",
            "regTime": "",
            "firstLevelStatus": 100000,
            "selfFirstLevelStatus": 100000,
            "nowDepartmentName": "花家地门店"
        },
        "assetInfoVo": {
            "bussLineId": 1,
            "bussLineTime": "",
            "willBussLineId": "",
            "willBussLineTime": "",
            "checkResultName": "",
            "allocationStatusName": "",
            "isInsideScrapName": "",
            "lastApproveSaleTime": "",
            "firstTransferOwnershipTime": "",
            "vehicleNoBeforeTransfer": ""
        },
        "financeInfoVo": {
            "carOwnerId": 29,
            "purchasePrice": "5000",
            "registrationAndPettyExpenses": "",
            "taxAmount": "",
            "incrementTaxExpenses": "1200",
            "monthDepreciationRate": "",
            "originalValue": "6200",
            "taxInvoiceDate": "",
            "purchaseDate": "2021-12-01",
            "remainderPrice": "",
            "vehicleAndVesselTax": "",
            "estimatedRemainderPrice": "",
            "realHandlePrice": "",
            "handlePettyExpenses": "",
            "realSellOutPrice": "",
            "handlePettyIncrementTax": "",
            "invoiceNo": ""
        },
        "garageInfoVo": {
            "lastTimeAttendMile": "",
            "nextTimeAttendMile": "",
            "vehicleFrom": ""
        },
        "deviceInfoVo": {
            "gps": 0,
            "deviceNo": "",
            "deviceType": ""
        },
        "noInfoPo": {
            "vehicleNoColor": ""
        },
        "colorPo": {
            "interiorColorId": ""
        },
        "takeoverPo": {
            "compulsoryInsurance": 1,
            "needInsured": 0
        },
        "insurancePos": [{
            "createTime": 1639374558566
        }],
        "modifyEmpName": null,
        "createEmpName": null,
        "createTime": null,
        "modifyTime": null
    }'''
    datapre=''' {	"baseInfoVo": {
		"vehicleId": "",
		"companyId": 46,
		"brandId": 67,
		"vehicleSeriesId": 560,
		"vehiclenoCityId": "",
		"vehiclenoCityName": "",
		"vehicleNo": "",
		"frameNo": "frameNo",
		"engineNo": "engineNo",
		"engineType": "",
		"envStandard": 3,
		"modelId": 1221,
		"vehicleLicensemode": "",
		"vehicleModel300Name": "2015款 宝马5系GT(进口) 528i 领先型",
		"color": 3,
		"registerNo": "",
		"useNature": 0,
		"carUseType": 0,
		"vehicleRemark": "",
		"checkTime": "",
		"ownerStatus1st": 10,
		"ownerStatus2nd": 1001,
		"ownerStatus3rd": "",
		"contractNo": "llypreht001"
	},
	"operatingInfoVo": {
		"shortModelId": 718,
		"shortModelName": "长租车型（特殊）",
		"businessType": 1,
		"isUCar": 0,
		"isHertz": 0,
		"isLease": 0,
		"onLineTime": "",
		"lastQuitRunTime": "",
		"handleTime": "",
		"ucarRentType": "",
		"oilVolume": null,
		"batteryVolume": null,
		"batteryPercentage": 0,
		"oilVolumeLatticeNum": 0,
		"orderCar": 0,
		"cityId": 1,
		"nowCityId": 1,
		"nowCityName": "北京",
		"parkCityId": 1,
		"departmentId": 75405,
		"nowDepartmentId": 75405,
		"parkDeptId": 75405,
		"runMilesInput": "100",
		"nextInspecteTime": "",
		"regTime": "",
		"firstLevelStatus": 100000,
		"selfFirstLevelStatus": 100000,
		"nowDepartmentName": "花家地门店"
	},
	"assetInfoVo": {
		"bussLineId": 1,
		"bussLineTime": "",
		"willBussLineId": "",
		"willBussLineTime": "",
		"checkResultName": "",
		"allocationStatusName": "",
		"isInsideScrapName": "",
		"lastApproveSaleTime": "",
		"firstTransferOwnershipTime": "",
		"vehicleNoBeforeTransfer": ""
	},
	"financeInfoVo": {
		"carOwnerId": 87,
		"purchasePrice": "150000",
		"registrationAndPettyExpenses": "",
		"taxAmount": "",
		"incrementTaxExpenses": "3000",
		"monthDepreciationRate": "",
		"originalValue": "153000",
		"taxInvoiceDate": "",
		"purchaseDate": "2021-12-02",
		"remainderPrice": "",
		"vehicleAndVesselTax": "",
		"estimatedRemainderPrice": "",
		"realHandlePrice": "",
		"handlePettyExpenses": "",
		"realSellOutPrice": "",
		"handlePettyIncrementTax": "",
		"invoiceNo": "2022012001"
	},
	"garageInfoVo": {
		"lastTimeAttendMile": "",
		"nextTimeAttendMile": "",
		"vehicleFrom": ""
	},
	"deviceInfoVo": {
		"gps": 0,
		"deviceNo": "",
		"deviceType": ""
	},
	"noInfoPo": {
		"vehicleNoColor": ""
	},
	"colorPo": {
		"interiorColorId": ""
	},
	"takeoverPo": {
		"compulsoryInsurance": 1,
		"needInsured": 1,
		"insuranceBeginDate": "2022-01-01",
		"insuranceEndDate": "2023-01-02"
	},
	"insurancePos": [{
		"createTime": 1642660788023,
		"insuranceTypeName": "第三者责任险",
		"needAmount": 0,
		"insuranceAmount": null
	}, {
		"createTime": 1642661000825,
		"insuranceTypeName": "车辆损失险",
		"needAmount": 1,
		"insuranceAmount": "100"
	}],
	"modifyEmpName": null,
	"createEmpName": null,
	"createTime": null,
	"modifyTime": null
}
    '''
    # print(type(data))
    mydata=json.loads(datapre)
    # print(mydata)
    # print(type(mydata))
    mydata["baseInfoVo"]["frameNo"]=frameNo
    mydata["baseInfoVo"]["engineNo"]=engineNo
    # print(mydata)
    # setcookie=login_do()
    # print(setcookie)
    head = {
    # "Cookie": "intranet-test-sessionid=BD8543E7FEC03FCCDD4CA86C2060E494",#每天执行的时候，cookie会发生变化，需要更新，该怎么处理啊？？？？？
    "Cookie":setcookie,
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36"
    }
    #当参数是json格式的时候，需要使用json格式传参，不能使用data
    re = requests.post("http://**", json=mydata, headers=head)
    print(re.text)

# createVehicle()
#读取Excel参数，并复制给参数--json
'''
python中使用openpyxl模块时报错： File is not a zip file。
最大的原因就是不是真正的 xlsx文件， 如果是通过 库xlwt  新建的文件，或者是通过自己修改后缀名得到的 xlsx文件，都会报错，
我遇到的解决办法基本都是自己使用 office 新建一个xlsx文件，网上说的是由于新版 office 加密的原因，只能通过 office 的软件才能建立真正的xlsx文件。
'''
#参数化方式批量创建车辆
def createcarsbyexcel():
    filepath="/Users/**.xlsx"
    #打开Excel文件
    xlsbook=openpyxl.load_workbook(filepath)
    #获取sheet页,当前获取的是第一个sheet
    booksheet=xlsbook.worksheets[0]
    '''
    print(sheet.max_row)#获取表格的最大行
    print(sheet.min_row)#获取表格的最小行
    print(sheet.max_column)#获取表格的最大列
    print(sheet.min_column)#获取表格的最小列
    '''
    print(booksheet.max_row,booksheet.max_column)
    setcookie = login_do()#登录后获取cookie并存储，供其他需要的接口使用
    print(setcookie)
    for i in range(2,booksheet.max_row+1):#行数
        for j in range(1,booksheet.max_column):#列数
            frameNo =booksheet.cell(i,j).value
            engineNo =booksheet.cell(i,j+1).value
            print(frameNo,engineNo)
            createVehicle(frameNo,engineNo,setcookie)


#随机产生车架号和发动机号方式批量创建车辆
def createcarsbyrandom():
    setcookie = login_do()#登录后获取cookie并存储，供其他需要的接口使用
    print(setcookie)
    for i in range(4,20):
        print(i)
        datetemp=str(datetime.date.today()).split("-")
        # print(datetemp, type(datetemp))
        datepre=datetemp[0]+datetemp[1]+datetemp[2]
        print(datepre)
        strpre="LLY"+datepre+"0"*(6-len(str(i)))+str(i)
        print(strpre)
        frameNo=strpre
        engineNo=strpre
        createVehicle(frameNo, engineNo, setcookie)

#createcarsbyexcel()

# createcarsbyrandom()


# list1=[]
# lenth=len(list1)
# print(lenth)

# url="https://api.oick.cn/dutang/api.php"
# response=requests.get(url)
# print(response.text)

#运营助手APP预生产环境登录
def APPlogin_do():
    data={
    "deviceId": "android_zuche_op_dff2c99b-50df-3dad-ae11-fa2eefe5a551",
    "deviceType": 1,
    "loginType": "1",
    "mobileModel": "Xiaomi MIX 3",
    "mobileSystem": "29",
    "passwd": "12345678",
    "username": "lly",
    "lat": "39.977249",
    "lon": "116.480575",
    "cid": "401100"
}


    # 需要用运营助手里面的接口日志传参，另外再加一个万能cid=912100.直接用抓包工具里面的传参会报“密钥过期”
    data1={
        # "cid": "401100",
        "cid": "912100",
        "event_id": "167",
        "q": "03QoAWjLVEXN5BEC4xJOTgxiQvqnRbSKp7y8IMruAyYHAmX-XvF51ryn6g-vb7bmV55Uy5KPrXPD1O6_CeFuHoGpARHD0DUBccc1PHRVXixNBrMcUVweFSBtkfZkW8hGGxkcGj8inv7oSduDI9Rn6tbBWwM41vsBcv7nW14QAGBsOpzn3CDIX7Z3baRAYeuylEcyb5e-x38ETH4F-7gHORoImp0ilxSak3VciwcR6YFJtKHsmI4xf_wEu764tN9UkVXb-nFs7RDam4HshSEjn4Ytz5SVHohviWj3gz_OTs7xHQtQ-ZA6dB0ywhnuoBDa-Wx8CRb5f0g5SZHCWS0kZR2BroiqoqJcfQEGuenFdQk=",
        "sign": "19410295963916857479842017351652221837",
        "uid": "8f39cbc5-9195-46c9-9faf-8764157e97bc1646208880867"
    }
    re=requests.post("http://**/v1",data=data)
    print(re.status_code)
    print(re.text)
    print(re.headers)
    # setcookie=re.headers['Set-Cookie'].split(";")[0]
    # print(re.headers['Set-Cookie'].split(";")[0])
    # return setcookie
# setcookie=APPlogin_do()
APPlogin_do()

