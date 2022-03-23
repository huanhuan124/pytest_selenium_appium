import requests
import re

from openpyxl import load_workbook

class Base_case:

    def request_test_session(self, url, method, data):
        session = requests.session()
        if method == 'GET':
            re = session.get(url, data=data)
            print("login接口返回的结果：")
            print(re.text)
            # print(re.headers)
            print(re.headers['Set-Cookie'])
            setcookie = re.headers['Set-Cookie'].split(";")[0]
            print(re.headers['Set-Cookie'].split(";")[0])
            return re
        elif method == 'POST':
            re = session.post(url, data=data)
            print("login接口返回的结果：")
            print(re.text)
            # print(re.headers)
            print(re.headers['Set-Cookie'])
            setcookie = re.headers['Set-Cookie'].split(";")[0]
            print(re.headers['Set-Cookie'].split(";")[0])
            return re

    def request_test(self,url,method,data ,header):

        if method == 'GET':
            re = requests.get(url, data = data,header=header)
            print("login接口返回的结果：")
            print(re.text)
            # print(re.headers)
            print(re.headers['Set-Cookie'])
            setcookie = re.headers['Set-Cookie'].split(";")[0]
            print(re.headers['Set-Cookie'].split(";")[0])
            return re
        elif method == 'POST':
            re = requests.post(url,data = data,header=header)
            print("login接口返回的结果：")
            print(re.text)
            # print(re.headers)
            print(re.headers['Set-Cookie'])
            setcookie = re.headers['Set-Cookie'].split(";")[0]
            print(re.headers['Set-Cookie'].split(";")[0])
            return  re

    def extract_data(maps, cls, json_res_str, json_res_dict):
        '''
        提取需要存储的数据
        @param maps:需要提取的映射{cls_arg:key}
        @param cls:测试用例类
        @param json_res_str:json响应字符串
        @param json_res_dict: json响应字典
        @return:
        '''
        # 提取C端应约时所需参数 arrPriceItems
        # if 'arrPriceItems' in maps.keys():
        #     temp = json_res_dict['data']['price']['arrPriceItems']
        #     temp[0]['selected'] = True
        #     setattr(cls, 'arrPriceItems', json.dumps(temp, ensure_ascii=False))
        #     #return让函数结束，不往下走
        #     return
        # 去掉干扰字符串
        json_res_str = json_res_str.replace(' ', "").replace("\n", "")
        for arg, key in maps.items():
            # 根据key生成正则表达式
            re_strs = [
                r'"{}":"(.*?)"'.format(key),  # 字符串参数
                r'"{}":(\d+\.\d+)'.format(key),  # 浮点数参数
                r'"{}":(\d+)'.format(key),  # 整数参数
                r'"{}":.(\w+).'.format(key)  # 提取[]参数
            ]

            # 取响应中的数据，并保存在类中
            for re_str in re_strs:
                value = re.findall(re_str, json_res_str)

                if value:
                    setattr(cls, arg, value[0])
                    # 找到跳出循环
                    break

    def load_cases(cls, sheet_name):
        '''
        获取excel中的用例数据
        :param sheet_name:表名
        :return:
        '''
        datafile = "D:\SoftwareInstall\pycharmworkspace\\testproject1228\普通功能\\testcase\data\cases.xlsx"
        return cls.get_data_from_excel(datafile, sheet_name)

    def get_data_from_excel(file_name, sheet_name=None):
        '''
        从excel读取数据
        :param file_name:文件名
        :param sheet_name:表单名
        :return:
        '''
        # 打开工作簿
        wb = load_workbook(file_name)
        if sheet_name is None:
            # 如果表单名为None获取当前表单
            sh = wb.active
        else:
            # 打开表单
            sh = wb[sheet_name]
        # 新建data存放excel中的数据
        data = []
        # 获取最大行最大列
        row = sh.max_row
        column = sh.max_column
        # 获取表头
        title = {}
        for i in range(1, column + 1):
            title[i] = sh.cell(row=1, column=i).value
        # 获取除表头以外的数据
        for j in range(2, row + 1):
            temp = {}
            for i in range(1, column + 1):
                temp[title[i]] = sh.cell(row=j, column=i).value
            data.append(temp)
        return data  # [{},{},{},{}]
